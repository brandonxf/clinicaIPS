"""
Modulo de Reportes - exportacion PDF, Excel, CSV
"""
import io, csv
from datetime import datetime
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from drf_spectacular.utils import extend_schema
from apps.authentication.permissions import EsMedicoOAdministrador, EsAnalistaOAdministrador, EsMedicoAnalistaOAdministrador
from apps.etl.models import Paciente, HistorialETL
from apps.etl.serializers import PacienteSerializer
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors


def _filtrar_pacientes(request):
    qs = Paciente.objects.all()
    riesgo  = request.query_params.get('riesgo')
    sexo    = request.query_params.get('sexo')
    critico = request.query_params.get('critico')
    busqueda = request.query_params.get('search')
    if riesgo:   qs = qs.filter(riesgo_enfermedad=riesgo)
    if sexo:     qs = qs.filter(sexo=sexo)
    if critico == 'true': qs = qs.filter(es_critico=True)
    if busqueda:
        from django.db.models import Q
        qs = qs.filter(Q(nombres__icontains=busqueda) | Q(apellidos__icontains=busqueda) | Q(id_paciente__icontains=busqueda))
    return qs


@extend_schema(
    tags=['reportes'],
    summary='Exportar pacientes como CSV',
    description='Descarga el listado completo de pacientes en formato CSV con BOM para Excel. Respeta filtros: riesgo, sexo, critico, search.',
)
@api_view(['GET'])
@permission_classes([IsAuthenticated, EsMedicoOAdministrador])
def exportar_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="pacientes.csv"'
    response.write('\ufeff')  # BOM para Excel
    writer = csv.writer(response)
    campos = ['id_paciente','nombres','apellidos','edad','sexo','peso','altura',
              'imc','clasificacion_imc','presion_sistolica','presion_diastolica',
              'glucosa','colesterol','saturacion_oxigeno','temperatura',
              'diagnostico_preliminar','riesgo_enfermedad','es_critico','fecha_consulta']
    writer.writerow(campos)
    qs = _filtrar_pacientes(request)
    for p in qs.iterator():
        writer.writerow([getattr(p, c, '') for c in campos])
    return response

@extend_schema(
    tags=['reportes'],
    summary='Exportar pacientes como Excel',
    description='Descarga el listado de pacientes en formato .xlsx con colores por nivel de riesgo. Respeta filtros: riesgo, sexo, critico, search.',
)
@api_view(['GET'])
@permission_classes([IsAuthenticated, EsMedicoOAdministrador])
def exportar_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from rest_framework.response import Response
        return Response({'error': 'openpyxl no instalado'}, status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pacientes"

    headers = ['ID','Nombres','Apellidos','Edad','Sexo','Peso','Altura','IMC',
               'Clasificación IMC','P. Sistólica','P. Diastólica','Glucosa',
               'Colesterol','Sat. Oxígeno','Temperatura','Diagnóstico','Riesgo','Crítico']
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    riesgo_colors = {'bajo': 'C6EFCE', 'medio': 'FFEB9C', 'alto': 'FFC7CE', 'critico': 'FF0000'}
    qs = _filtrar_pacientes(request)
    for row_num, p in enumerate(qs.iterator(), 2):
        row = [p.id_paciente, p.nombres, p.apellidos, p.edad, p.sexo,
               p.peso, p.altura, p.imc, p.clasificacion_imc,
               p.presion_sistolica, p.presion_diastolica, p.glucosa,
               p.colesterol, p.saturacion_oxigeno, p.temperatura,
               p.diagnostico_preliminar, p.riesgo_enfermedad, 'Sí' if p.es_critico else 'No']
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
        riesgo = p.riesgo_enfermedad or 'bajo'
        color = riesgo_colors.get(riesgo, 'FFFFFF')
        ws.cell(row=row_num, column=17).fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_pacientes.xlsx"'
    return response

@extend_schema(
    tags=['reportes'],
    summary='Historial ETL como reporte',
    description='Retorna las ultimas 50 ejecuciones del proceso ETL en formato JSON.',
)
@api_view(['GET'])
@permission_classes([IsAuthenticated, EsAnalistaOAdministrador])
def historial_etl_reporte(request):
    from rest_framework.response import Response
    from apps.etl.serializers import HistorialETLSerializer
    data = HistorialETLSerializer(HistorialETL.objects.all()[:50], many=True).data
    return Response(data)

@extend_schema(
    tags=['reportes'],
    summary='Exportar reporte clinico en PDF',
    description='Genera un PDF con resumen de KPIs y listado completo de pacientes. Respeta filtros: riesgo, sexo, critico, search.',
)
@api_view(['GET'])
@permission_classes([IsAuthenticated, EsMedicoAnalistaOAdministrador])
def exportar_pdf(request):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(name='DocTitle', parent=styles['Heading1'], fontSize=20, leading=24, textColor=colors.HexColor('#1F4E79'), spaceAfter=6)
    meta_style = ParagraphStyle(name='DocMeta', parent=styles['Normal'], fontSize=9, leading=12, textColor=colors.HexColor('#555555'), spaceAfter=15)
    section_title_style = ParagraphStyle(name='SectionTitle', parent=styles['Heading2'], fontSize=12, leading=16, textColor=colors.HexColor('#1F4E79'), spaceBefore=15, spaceAfter=8)
    cell_style = ParagraphStyle(name='TableCell', parent=styles['Normal'], fontSize=8, leading=10, textColor=colors.HexColor('#333333'))
    header_style = ParagraphStyle(name='TableHeader', parent=styles['Normal'], fontSize=8, leading=10, textColor=colors.white, fontName='Helvetica-Bold')

    qs = _filtrar_pacientes(request)
    total = qs.count()
    criticos = qs.filter(es_critico=True).count()
    hipertensos = qs.filter(presion_sistolica__gt=140).count()
    diabeticos = qs.filter(glucosa__gt=126).count()

    story.append(Paragraph("HealthAnalytics IPS — Plataforma de Analítica Clínica", title_style))
    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
    story.append(Paragraph(f"Reporte Clínico de Pacientes · Generado el: {fecha_actual}", meta_style))
    story.append(Spacer(1, 5))

    story.append(Paragraph("Resumen de Estadísticas Clínicas", section_title_style))
    resumen = Table([
        [Paragraph("<b>Total Pacientes:</b>", cell_style), Paragraph(str(total), cell_style),
         Paragraph("<b>Críticos:</b>", cell_style), Paragraph(f"<font color='red'><b>{criticos}</b></font>", cell_style)],
        [Paragraph("<b>Hipertensos (Sist. > 140):</b>", cell_style), Paragraph(str(hipertensos), cell_style),
         Paragraph("<b>Diabéticos (Glucosa > 126):</b>", cell_style), Paragraph(str(diabeticos), cell_style)],
    ], colWidths=[150, 100, 150, 100])
    resumen.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F4F6FB')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#D2D7E5')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E7F0')),
        ('PADDING', (0,0), (-1,-1), 6), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(resumen)
    story.append(Spacer(1, 15))

    filtros_aplicados = []
    if request.query_params.get('riesgo'): filtros_aplicados.append(f"Riesgo: {request.query_params['riesgo']}")
    if request.query_params.get('sexo'):   filtros_aplicados.append(f"Sexo: {request.query_params['sexo']}")
    if request.query_params.get('critico') == 'true': filtros_aplicados.append("Solo críticos")
    if request.query_params.get('search'): filtros_aplicados.append(f"Búsqueda: {request.query_params['search']}")
    subtitulo = "Listado Completo de Pacientes"
    if filtros_aplicados:
        subtitulo += " — Filtros: " + ", ".join(filtros_aplicados)
    story.append(Paragraph(subtitulo, section_title_style))

    if total == 0:
        story.append(Paragraph("No se encontraron pacientes con los filtros seleccionados.", cell_style))
    else:
        table_data = [[
            Paragraph("ID", header_style), Paragraph("Nombre Completo", header_style),
            Paragraph("Edad", header_style), Paragraph("Sexo", header_style),
            Paragraph("IMC", header_style), Paragraph("Glucosa", header_style),
            Paragraph("P. Sistólica", header_style), Paragraph("Riesgo", header_style),
        ]]
        for p in qs.iterator():
            nombre = f"{p.nombres} {p.apellidos}"
            color = "#FF0000" if p.riesgo_enfermedad == 'critico' else ("#FF8000" if p.riesgo_enfermedad == 'alto' else ("#E6B800" if p.riesgo_enfermedad == 'medio' else "#059669"))
            riesgo_p = Paragraph(f"<font color='{color}'><b>{p.riesgo_enfermedad.upper() if p.riesgo_enfermedad else '—'}</b></font>", cell_style)
            table_data.append([
                Paragraph(str(p.id_paciente), cell_style), Paragraph(nombre, cell_style),
                Paragraph(str(p.edad or '—'), cell_style), Paragraph(str(p.sexo or '—'), cell_style),
                Paragraph(f"{p.imc:.1f}" if p.imc else '—', cell_style),
                Paragraph(str(p.glucosa or '—'), cell_style),
                Paragraph(str(p.presion_sistolica or '—'), cell_style), riesgo_p,
            ])
        pacientes_table = Table(table_data, colWidths=[40, 140, 35, 35, 45, 50, 55, 60])
        pacientes_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
            ('PADDING', (0,0), (-1,-1), 4),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9F9F9')]),
        ]))
        story.append(pacientes_table)

    doc.build(story)
    buffer.seek(0)
    pdf_data = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_analitica_clinica.pdf"'
    return response
